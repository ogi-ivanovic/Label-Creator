import os
import openpyxl
#'C:/Users/Ogi/Desktop/test.xlsx'

def formatDate(date):
    year = date[0:4]
    month = date[5:7]
    day = date[8:10]

    if day[0] == '0':
        day = day[-1]

    if month == '01':
        month = 'January'
    if month == '02':
        month = 'February'
    if month == '03':
        month = 'March'
    if month == '04':
        month = 'April'
    if month == '05':
        month = 'May'
    if month == '06':
        month = 'June'
    if month == '07':
        month = 'July'
    if month == '08':
        month = 'August'
    if month == '09':
        month = 'September'
    if month == '10':
        month = 'October'
    if month == '11':
        month = 'November'
    if month == '12':
        month = 'December'

    formattedDate = month + ' ' + day + ', ' + year
    return formattedDate

def combinedFormat(c):
    if len(c) < 61:
        return [c, '']
    i = 60

    while True:
        if c[i] == ' ':
            break
        i -= 1

    return [c[:i], c[i+1:]]


try:
    os.remove('R:/ContEd/Classroom/Classroom Scheduling Labels/spreadsheet for mail merge.xlsx')
except OSError:
    pass

CELPath = 'R:/ContEd/Classroom/Classroom Scheduling Labels/schedule for labels.xlsx'
myPath = 'example.xlsx'

dataOriginal = openpyxl.load_workbook(CELPath)
dataSheetOriginal = dataOriginal.active #sheet with all the original course info

data = openpyxl.Workbook()
dataSheet = data['Sheet'] #sheet with all the info from the original we need

maxOriginalRow = dataSheetOriginal.max_row

dataDateCol = 1
dataTitleCol = 3
dataRoomCol = 6
currDataRow = 1
for row in range(2, maxOriginalRow + 1):
    room = dataSheetOriginal.cell(row=row, column=dataRoomCol).value

    if room != None:
        currDataRow += 1

        dataDate = dataSheetOriginal.cell(row=row, column=dataDateCol).value
        dataSheet.cell(row=currDataRow, column=1).value = dataDate

        dataTitle = dataSheetOriginal.cell(row=row, column=dataTitleCol).value
        dataSheet.cell(row=currDataRow, column=2).value = dataTitle

        dataRoom = dataSheetOriginal.cell(row=row, column=dataRoomCol).value
        dataSheet.cell(row=currDataRow, column=3).value = dataRoom


wb = openpyxl.Workbook()
sheet = wb['Sheet'] #sheet I will be writing to

maxRow = dataSheet.max_row

courseCol = 2
classroomCol = 3
combinedCol = 2
dateCol = 1
tempCombinedCol1 = 4
tempCombinedCol2 = 5

for row in range(2, maxRow + 1):
    course = str(dataSheet.cell(row=row, column=courseCol).value)
    classroom = str(dataSheet.cell(row=row, column=classroomCol).value)
    combined = course + ' - ' + classroom
    combined = combinedFormat(combined)
    combined1 = combined[0]
    combined2 = combined[1]

    dataSheet.cell(row=row, column=tempCombinedCol1).value = combined1
    dataSheet.cell(row=row, column=tempCombinedCol2).value = combined2

sheet['A1'] = 'Date'
sheet['B1'] = 'Course 1 - 1'
sheet['C1'] = 'Course 1 - 2'
sheet['D1'] = 'Course 2 - 1'
sheet['E1'] = 'Course 2 - 2'
sheet['F1'] = 'Course 3 - 1'
sheet['G1'] = 'Course 3 - 2'

currClassCol = 2
sheetRow = 2
lastDate = formatDate(str(dataSheet['A2'].value)[:-8])
sheet['A2'].value = lastDate
sheet['B2'].value = dataSheet['D2'].value
sheet['C2'].value = dataSheet['E2'].value

for dataSheetRow in range(3, maxRow + 1):
    date = str(dataSheet.cell(row=dataSheetRow, column=dateCol).value)[:-8]
    date = formatDate(date)

    if date == lastDate:
        currClassCol += 2
    else:
        currClassCol = 2
        sheetRow += 1
        lastDate = date
        sheet.cell(row=sheetRow, column=dateCol).value = date

    value1 = dataSheet.cell(row=dataSheetRow, column=tempCombinedCol1).value
    value2 = dataSheet.cell(row=dataSheetRow, column=tempCombinedCol2).value
    sheet.cell(row=sheetRow, column=currClassCol).value = value1
    sheet.cell(row=sheetRow, column=currClassCol+1).value = value2


CELPath = 'R:/ContEd/Classroom/Classroom Scheduling Labels/spreadsheet for mail merge.xlsx'
myPath = 'example_copy.xlsx'
wb.save(CELPath)

try:
    os.remove('R:/ContEd/Classroom/Classroom Scheduling Labels/schedule.xlsx')
except OSError:
    pass
